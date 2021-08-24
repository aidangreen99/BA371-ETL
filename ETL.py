import os
import openpyxl
import sys
import getopt
import sqlite3

#py ETL.py "C:\Users\Aidan\Documents\BA371\Group\testdb - Copy.SQLITE" "C:\Users\Aidan\Documents\BA371\Group\research_productivity\research_productivity"

#setting up global variables
role_dict = {}
activity_type_dict = {}
target_type_dict = {}
dept_dict = {}
db_file = ''
spreadsheet_root = ''

#Necessary lists
#targets_data = []
#target_types_data = []
#co_auth_data = []
depts_list = ["Accounting", "Marketing", "Finance", "BIS", "Management", "Entrepreneurship"]
target_types_list = ["journal", "conference"]
activity_types_list = ["submitted", "accepted", "r&r", "rejected"]
roles_list = ["Contributor", "Lead", "Co_lead"]
#Not the best solution for faculty list, see readme for explanation
fac_list = ["Josie Ross", "Tallulah Stewart", "Bridget Cox", "Sally Miller", "Bella Foster", 
          "Sapphire Carter", "Amelia Lewis", "Mindy Garcia", "Ebony Taylor", "Carrie Perry", 
          "Anaya Kelly", "Ayat Reed", "Ria Green", "Klaudia Jenkins", "Myah Watson", 
          "Alma Hughes", "Hannah Turner", "Henna Baker", "May Jackson", "Jo Barnes", 
          "Tayla Sanchez", "Bailey Howard", "Louis Bennett", "Caitlan Butler", "Darcey Davis", 
          "Alfred White", "Shayne Thomas", "Melvin Richardson", "Marvin Cook", "Leyton Ward", 
          "Jesse Wright", "Barney Torres", "Casper Brown", "Aarav Roberts", "Alexander Morgan", 
          "Wesley Adams", "Carwyn Rogers", "Simone Clark", "Wilbur Parker", "Kinsley Thompson",
          "Zakaria Cooper", "Garfield Anderson", "Pierce Bailey", "Dan Allen", "Rick Wilson",
          "Colin Jones", "Ayub Myers", "Robert Brooks", "Darcy Sullivan", "Howard Jones"]

class paper_data:
     "This helps keep track of what belongs to what paper"
     def __init__(self):
          self.paper_title = ""
          self.coauthors = []
          self.target = ""
          self.target_type = ""
          self.fac_role = ""
          self.activity_dates = []
          self.activity = []




class sheetData:
     "This class allows easy storage and retreival of data per faculty"
     def __init__(self):
          self.faculty_name = ""
          self.dept_name = ""
          self.dept_id = 999
          self.papers = []
     




def clear_transaction_tables():
  target_query = "delete from target;"
  authors_query = "delete from fac_paper;"
  coauthors_query = "delete from co_auth_paper;"
  papers_query = "delete from papers;"
  activities_query = "delete from activities;"
  try:
    cursor.execute(target_query)
    cursor.execute(authors_query)
    cursor.execute(coauthors_query)
    cursor.execute(activities_query)
    cursor.execute(papers_query)
  except Exception as err:
    print("Error clearing out transaction tables...\n", err)
    exit(1)
  connection.commit()

#Fill the depts table, specifically to generate an ID for dict matching
def dept_fill():
     cursor.execute("DELETE FROM depts;")
     for dep in depts_list:
          query = "INSERT INTO depts (dept_name) VALUES ('"
          query += dep
          query += "');"
          #print(query)
          try:
               cursor.execute(query)       
          except Exception as err:
               print("QUERY EXECUTION ERROR: " + str(err))
               break
     connection.commit()        

#Fill the target_type table, specifically to generate an ID for dict matching
def target_types_fill():
     cursor.execute("DELETE FROM target_type;")
     for target_type in target_types_list:
          query = "INSERT INTO target_type (target_type_name) VALUES ('"
          query += target_type
          query += "');"
          #print(query)
          try:
               cursor.execute(query)       
          except Exception as err:
               print("QUERY EXECUTION ERROR: " + str(err))
               break
     connection.commit()        

#Fill the activity_type table, specifically to generate an ID for dict matching
def activity_types_fill():
     cursor.execute("DELETE FROM activity_type;")
     for activity_type in activity_types_list:
          query = "INSERT INTO activity_type (activity_type) VALUES ('"
          query += activity_type
          query += "');"
          #print(query)
          try:
               cursor.execute(query)       
          except Exception as err:
               print("QUERY EXECUTION ERROR: " + str(err))
               break
     connection.commit()        

#Fill the roles table, specifically to generate an ID for dict matching
def roles_fill():
     cursor.execute("DELETE FROM roles;")
     for role in roles_list:
          query = "INSERT INTO roles (role) VALUES ('"
          query += role
          query += "');"
          #print(query)
          try:
               cursor.execute(query)       
          except Exception as err:
               print("QUERY EXECUTION ERROR: " + str(err))
               break
     connection.commit()     

#Fill the faculty table, specifically to generate an ID for dict matching
def faculty_fill():
     cursor.execute("DELETE FROM faculty;")
     for name in fac_list:
          query = "INSERT INTO faculty (faculty_name) VALUES ('"
          query += name
          query += "');"
          #print(query)
          try:
               cursor.execute(query)       
          except Exception as err:
               print("QUERY EXECUTION ERROR: " + str(err))
               break
     connection.commit()     

def dict_fill():
     #Set up the look-up dicts for role_type, activity_type and target_type ids
     #Role types:
     query = "select role_id, role from roles;"
     try:
          cursor.execute(query)
     except Exception as err:
          print("Error executing query...\n", err)
          exit(1)
     for record in cursor.fetchall():
          role_dict[record[1]] = record[0]
          #print(record)

     #Activity types:
     query = "select activity_type_id, activity_type from activity_type;"
     try:
          cursor.execute(query)
     except Exception as err:
          print("Error executing query...\n", err)
          exit(1)
     for record in cursor.fetchall():
          activity_type_dict[record[1]] = record[0]
          #print(record)

     #Target types:
     query = "select target_type_id, target_type_name from target_type;"
     try:
          cursor.execute(query)
     except Exception as err:
          print("Error executing query...\n", err)
          exit(1)
     for record in cursor.fetchall():
          target_type_dict[record[1]] = record[0]
          #print(record)

     #Department IDs:
     query = "select dept_id, dept_name from depts;"
     try:
          cursor.execute(query)
     except Exception as err:
          print("Error executing query...\n", err)
          exit(1)
     for record in cursor.fetchall():
          dept_dict[record[1]] = record[0]
          #print(record)

def processsheets():
     #Variable 'spreadsheet_root' holds the name of the folder containing all the faculty folders
     for folder in os.listdir(spreadsheet_root):
          excel_file = spreadsheet_root + "\\" + folder + "\\" + folder + ".xlsx"
          #print("\n" + "Working on file: " + folder)
          try:
               wb = openpyxl.load_workbook(excel_file, data_only=True)
          except Exception as err:
               print("Error opening Excel file: " + excel_file + "\n", err)
               exit(1)

          #Get the sheet from the current workbook
          sheet = wb.worksheets[0]
          current_process = sheetData()
          paper_index = 0
          if sheet.max_row > 6:
               #Process the spreadsheet code goes here
               current_process = sheetData()
               current_process.faculty_name = sheet['A3'].value
               current_process.dept_name = sheet['B3'].value
               current_process.dept_id = dept_dict[sheet['B3'].value]
               current_process.papers.clear()
               for row in sheet.iter_rows(min_row=7, max_col=1, max_row=sheet.max_row):
                    for cell in row:
                         if cell.value != None:
                              current_process.papers.append(paper_data())
                              current_process.papers[paper_index].paper_title = cell.value
                              current_process.papers[paper_index].target_type = cell.offset(row=0,column=1).value
                              current_process.papers[paper_index].target = cell.offset(row=0,column=2).value
                              current_process.papers[paper_index].coauthors.clear()
                              #if cell.offset(row=0,column=2).value not in targets_data: targets_data.append(cell.offset(row=0,column=2).value)
                              for num, coauth in enumerate(range(4)):
                                   if cell.offset(row=0, column = (num + 4)).value != None: current_process.papers[paper_index].coauthors.append(cell.offset(row=0, column = (num + 4)).value)
                              current_process.papers[paper_index].fac_role = cell.offset(row=0, column = 9).value
                              current_process.papers[paper_index].activity_dates.append(cell.offset(row=0, column = 10).value.strftime('%Y-%m-%d'))
                              current_process.papers[paper_index].activity.append(cell.offset(row=0, column = 11).value)
                              paper_index += 1
                         elif cell.value == None:
                              current_process.papers[paper_index-1].activity_dates.append(cell.offset(row=0, column = 10).value.strftime('%Y-%m-%d'))
                              current_process.papers[paper_index-1].activity.append(cell.offset(row=0, column = 11).value)
                    
               for paper in current_process.papers:
                    target_query = "insert into target (target_name, target_type_id) VALUES ('" + paper.target + "', " + str(target_type_dict[paper.target_type]) + ");"          
                    try:
                         cursor.execute(target_query)
                    except Exception as err:
                         pass
                         #print("Error inserting into target: " + str(err) + "\n (this isn't necessarily a bad thing)")
                    #connection.commit()
                    paper_query = "insert into papers (title, target_id) VALUES ('" + paper.paper_title + "', (select target_id from target where target_name = '" + paper.target + "'));"
                    try:
                         cursor.execute(paper_query)
                    except Exception as err:
                         pass
                         #print("Error inserting into papers: " + str(err))
                    #connection.commit()
                    fac_paper_query = "insert into fac_paper (faculty_id, paper_id, role_id) VALUES ((select faculty_id from faculty where faculty_name='" + current_process.faculty_name + "'), (select paper_id from papers where title='" + paper.paper_title + "'), " + str(role_dict[paper.fac_role]) + ");"
                    try:
                         cursor.execute(fac_paper_query)
                    except Exception as err:
                         print("Error inserting into fac_paper: " + str(err))
                    #connection.commit()
                    for itemID, item in enumerate(paper.activity):
                         activity_query = "insert into activities (activity_type_id, activity_date, paper_id) VALUES (" + str(activity_type_dict[item]) + ", '" + str(paper.activity_dates[itemID]) + "', (select paper_id from papers where title='" + paper.paper_title + "'));"
                         try:
                              cursor.execute(activity_query)
                         except Exception as err:
                              print("Error inserting into activities: " + str(err))
                    #connection.commit()
                    for item in paper.coauthors:
                         co_auth_query = " insert into co_auth_paper (co_auth_name, paper_id) VALUES ('" + item + "', (select paper_id from papers where title='" + paper.paper_title + "'));"
                         try:
                              cursor.execute(co_auth_query)
                         except Exception as err:
                              pass
                              #print("Error with query: " + co_auth_query + ", " + str(err))
                    #connection.commit()
                    fac_dep_query = "insert into fac_dept (faculty_id, dept_id) values ((select faculty_id from faculty where faculty_name = '" + current_process.faculty_name +"'), " + str(current_process.dept_id) + ");"         
                    try:
                         cursor.execute(fac_dep_query)
                    except Exception as err:
                         pass
                         #print("Error inserting into fac_dep: " + str(err))     
                    connection.commit()

          wb.close()





#Checking command line arguments
if len(sys.argv) != 3:
     print("etl.py \"<database_file_path>\" \"<spreadsheet_root_path>\"")
     sys.exit()
     
#assigning DB and spreadsheet file paths
db_file = sys.argv[1]
spreadsheet_root = sys.argv[2]


#connecting to DB
try:
     connection = sqlite3.connect(db_file)
except Exception as err:
     print("CONNECTION ERROR: " + str(err))
cursor =  connection.cursor()


#main basically
clear_transaction_tables()
dept_fill()
target_types_fill()
activity_types_fill()
roles_fill()
faculty_fill()
dict_fill()
processsheets()


cursor.close()
connection.close()
sys.exit()
